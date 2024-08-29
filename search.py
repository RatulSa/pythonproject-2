from tkinter import *
from tkinter import ttk
import openpyxl,os,tempfile
import tkinter.messagebox as tmsg

def search_window():
    root2 = Toplevel()
    root2.config(bg='#326273')
    root2.geometry('600x700')
    root2.title('Search')

    def clear():
          text_area.delete('1.0',END)
          name.set('')
          deptVar.set('Select Department')
          uid.set(0)
          detail_frame.update()

    def save():
          if text_area.get(1.0,END) == '\n':
                tmsg.showwarning('WARNING','Fill out details before saving')
          else:
                file = tempfile.mktemp('.txt')
                open(file,'w').write(text_area.get(1.0,END))
                os.startfile(file,'print')

    def search():
        if name.get()=='' and deptVar.get()=='Select Department' and uid.get() == 0 :
            tmsg.showerror('Error','Please Fill all Details')
        else:
            department = deptVar.get()+".xlsx"
            # filepath =  "C:\\Users\\ratul\\Desktop\\python gui\\pythonproject 2\\"+department
            filepath = department
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            maxcol = sheet.max_column
            maxrow = sheet.max_row
            row_data = []
            for i in range(1,maxrow+1):
                    
                    if sheet.cell(row=i,column=2).value == uid.get():
                            for j in range(1,maxcol+1):
                                    row_data.append(sheet.cell(row=i,column=j).value)
            # print(row_data)
            text_area.insert(END,"     St Thomas College of Engineering & Technology\n")
            text_area.insert(END,"=======================================================\n")
            text_area.insert(END,"                    APLICATION FORM                    \n\n")
            text_area.insert(END,"------Student Details----------------------------------\n\n")
            text_area.insert(END,f"Full Name : {row_data[0]}  \n")
            text_area.insert(END,f"Application no. : {row_data[1]}\n")
            text_area.insert(END,f"Department : {deptVar.get()}  \n")
            text_area.insert(END,f"Phone : {row_data[6]}\n")
            text_area.insert(END,f"Email : {row_data[5]}  \n\n")
            text_area.insert(END,"------Education----------------------------------------\n\n")
            text_area.insert(END,f"Higher Secondary Board : {row_data[14]}  \n")
            text_area.insert(END,f"12th marks(%) : {row_data[16]}\n")
            text_area.insert(END,f"Secondary Board : {row_data[15]}         \n")
            text_area.insert(END,f"10th marks(%) : {row_data[17]}\n\n")
            text_area.insert(END,"------Adress-------------------------------------------\n\n")
            text_area.insert(END,f"State : {row_data[10]}      P.O : {row_data[11]} \n")
            text_area.insert(END,f"P.S : {row_data[12]}         Pin : {row_data[13]} \n\n")
            text_area.insert(END,"------Parent Details-----------------------------------\n\n")
            text_area.insert(END,f"Father : {row_data[7]}      Mother : {row_data[8]} \n")
            text_area.insert(END,f"Phone  : {row_data[9]}   \n")
    name = StringVar()
    uid = IntVar()
    deptVar = StringVar()
    detail_frame = LabelFrame(root2,text='Fill Details',font='lucida 19 bold',borderwidth=5, relief='raised',fg="white",bg="grey20")
    detail_frame.pack(pady=20,)

    name_label = Label(detail_frame,text='Name',font='Bahnschrift',fg="white",bg="grey20")
    name_label.grid(row=0,column=0,sticky="w")
    name_entry = Entry(detail_frame,font='lucida 15 bold',textvariable=name)
    name_entry.grid(row=0,column=1,pady=5,)

    application_no_label = Label(detail_frame,text='Application No.',font='Bahnschrift',fg="white",bg="grey20")
    application_no_label.grid(row=1,column=0)
    application_no = Entry(detail_frame,font='lucida 15 bold',textvariable=uid)
    application_no.grid(row=1,column=1,pady=5)

    dept = [
        "Computer Science & Engineering","Electronics & Communication Engineering","Electrical Engineering","Artificial Intelligence & Machine Learinging","Information Technology"
    ]
     #'deptVar' is a string variable to store Department name
    deptVar.set("Select Department")
    dept_no_label = Label(detail_frame,text='Department',font='Bahnschrift',fg="white",bg="grey20")
    dept_no_label.grid(row=2,column=0,sticky="w")
    dept = ttk.Combobox(detail_frame,values=dept,textvariable=deptVar,font='lucida 10').grid(row=2,column=1,pady=5)

    button_frame = Frame(root2,borderwidth=3,relief=SUNKEN)
    button_frame.pack()

    search_btn = Button(button_frame,text="View",padx = 5,bg="grey",fg="white",font="lucida 15 bold",command=search)
    search_btn.grid(row=0,column=0,)

    save_btn = Button(button_frame,text = "Save",padx = 5,bg="grey",fg="white",font="lucida 15 bold",command=save)
    save_btn.grid(row=0,column=1,)

    clear_btn = Button(button_frame,text = "Clear",padx = 5,bg="grey",fg="white",font="lucida 15 bold",command=clear)
    clear_btn.grid(row=0,column=2,)
    text_area  = Text(root2,borderwidth=6,relief=GROOVE,width=55)
    text_area.pack(pady=30)

    root2.mainloop()